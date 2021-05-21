import pandas as pd
from sqlalchemy import create_engine
import collections
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os


def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            collen = df[col].apply(lambda x: len(str(x).encode())).max()
            ws.column_dimensions[letter].width = collen * \
                1.2 + 4  # 也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)

class UnqualHistoryDispatcher:

    def __init__(self):
        self.conn = create_engine(
            'mysql+pymysql://remote_root:symphony@170.0.35.150:3306/unqual')
        self.resDir = "files/"

    def modifyExcelFormat(self, filePath):
        wb = load_workbook(filePath)
        wb = self.changeColumnWidth(filePath, wb)
        wb = self.changeColumnColor(filePath, wb)
        wb.save(filePath)

    def changeColumnWidth(self, filePath, wb):
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            df = pd.read_excel(filePath, sheet).fillna('-')
            df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
            for col in df.columns:
                index = list(df.columns).index(col)  # 列序号
                letter = get_column_letter(index + 1)  # 列字母
                # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
                collen = df[col].apply(lambda x: len(str(x).encode())).mean()
                # 也就是列宽为最大长度*1.2 可以自己调整

                if collen < len(col):
                    finalWidth = len(col) * 1.1 + 10
                else:
                    finalWidth = collen * 1.1 + 2
                if col == "缺陷详细信息":
                    finalWidth = 20

                ws.column_dimensions[letter].width = finalWidth
        return wb

    def changeColumnColor(self, filePath, wb):

        redFill = PatternFill(fill_type='solid', fgColor="FE4365")
        orangeFill = PatternFill(fill_type='solid', fgColor="FC9D9A")
        yellowFill = PatternFill(fill_type='solid', fgColor="F9CDAD")

        reasonColumnMap = collections.OrderedDict()
        reasonColumnMap["技术员分析原因"] = redFill
        reasonColumnMap["FSP平台分析原因"] = orangeFill
        reasonColumnMap["TMEIC分析原因"] = yellowFill

        sheetName = "Sheet1"
        df = pd.read_excel(filePath, sheetName).fillna('-')

        for reason, colorFill in reasonColumnMap.items():
            index = list(df.columns).index(reason)
            letter = get_column_letter(index + 1)

            ws = wb[sheetName]
            ws.column_dimensions[letter].fill = colorFill
        return wb

    def trasformDataFrame(self, df):
        df['unqual_type_tag'] = (
            df['unqual_source'].apply(
                lambda x: "[{}]".format(x)) + df['unqual_type']
        )
        df['unqual_value_tag'] = (
            df['unqual_source'].apply(
                lambda x: "[{}]".format(x)) + df['unqual_value']
        )
        res = pd.DataFrame()
        g = df.groupby('coil_id')
        res['coil_id'] = list(g.groups.keys())
        print(res)
        for coilAttr in self.getCoilAttrColumns()[1:]:
            res[coilAttr] = list(g[coilAttr].first())
        res['unqual_category'] = list(g['unqual_category'].first())

        coilIds = res['coil_id']
        for i, coilId in enumerate(coilIds):
            df_coil = df.loc[df['coil_id'] == coilId]

            for subType in self.getUnqualSubTypeColumns():
                res.loc[i, subType] = ""

            self.addSubType(res, i, df_coil)

        return res

    def addSubType(self, res, idx, df):
        res.loc[idx, "unqual_value"] = ""
        for i in df.index:
            record = df.loc[i]
            unqualSource = record["unqual_source"]
            subType = self.getUnqualSourceToSubTypeMap()[unqualSource]
            print(res.loc[idx])
            res.loc[idx, subType] += record['unqual_type_tag']
            res.loc[idx, "unqual_value"] += record['unqual_value_tag']

    def getUnqualSourceToSubTypeMap(self):
        m = collections.OrderedDict()
        m["二级"] = "high_accu_type"
        m["FSP"] = "internal_ctrl_type"
        m["QMS"] = "internal_ctrl_type"
        m["质检"] = "shipped_type"
        return m

    def getColumnMap(self):
        m = collections.OrderedDict()
        m["coil_id"] = "热卷号"
        m["product_time"] = "生产时间"
        m["steel_grade"] = "钢种"
        m["aim_thk"] = "厚度"
        m["aim_wid"] = "宽度"
        m["unqual_category"] = "缺陷分类"
        m["high_accu_type"] = "高精度缺陷类型"
        m["internal_ctrl_type"] = "内控缺陷类型"
        m["shipped_type"] = "交货缺陷类型"
        m["unqual_value"] = "缺陷详细信息"
        m["technician_reason"] = "技术员分析原因"
        m["fsp_reason"] = "FSP平台分析原因"
        m["tmeic_reason"] = "TMEIC分析原因"
        m["root_reason"] = "根本原因"
        return m

    def getCoilAttrColumns(self):
        return list(self.getCoilAttrColumnMap().keys())

    def getCoilAttrColumnMap(self):
        m = collections.OrderedDict()
        m["coil_id"] = "热卷号"
        m["product_time"] = "生产时间"
        m["steel_grade"] = "钢种"
        m["aim_thk"] = "厚度"
        m["aim_wid"] = "宽度"
        return m

    def getUnqualSubTypeColumns(self):
        return self.getUnqualSubTypeColumnMap().keys()

    def getUnqualSubTypeColumnMap(self):
        m = collections.OrderedDict()
        m["high_accu_type"] = "高精度缺陷类型"
        m["internal_ctrl_type"] = "内控缺陷类型"
        m["shipped_type"] = "交货缺陷类型"
        return m

    def getReasonColumns(self):
        return self.getReasonColumnMap().keys()

    def getReasonColumnMap(self):
        m = collections.OrderedDict()
        m["technician_reason"] = "技术员分析原因"
        m["fsp_reason"] = "FSP平台分析原因"
        m["tmeic_reason"] = "TMEIC分析原因"
        m["root_reason"] = "根本原因"
        return m

    def changeColumnNames(self, df):
        res = pd.DataFrame()
        for k, v in self.getColumnMap().items():
            if k in df.columns:
                res[v] = df[k]
            else:
                res[v] = ""
        return res

    def getResultFilePath(self, serialNum, unqualCategory, startDate, endDate):
        curDir = self.resDir + "{}_{}/".format(startDate, endDate)
        os.makedirs(curDir, exist_ok=True)

        filePath = curDir + "{}_{}_不合格台账_{}_{}.xlsx".format(
            serialNum, unqualCategory, startDate, endDate
        )
        return filePath

    def dispatchUnqualHistoryByDate(self, currentDate):
        df = self.readDataByDate(currentDate)
        unqualCategories = self.getUnqualCategory()

        for i, unqualCategory in enumerate(unqualCategories):
            serialNum = "{0:02d}".format(i + 1)
            print(serialNum, unqualCategory)
            res = self.getUnqualHistoryDataByCategory(df, unqualCategory)
            res = self.trasformDataFrame(res)

            res = self.changeColumnNames(res)
            filePath = self.getResultFilePath(
                serialNum, unqualCategory, currentDate, currentDate
            )
            res.to_excel(filePath, index=False)
            self.modifyExcelFormat(filePath)
            print(serialNum, unqualCategory, "COMPLETED")

    def getUnqualHistoryDataByCategory(self, df, category):
        res = df.loc[df['unqual_category'] == category]
        res = res.reset_index(drop=True)
        return res

    def getUnqualCategory(self):
        return [
            "厚度",
            "宽度",
            "终轧温度",
            "卷取温度",
            "楔形",
            "凸度",
            "对称平直度",
            "非对称平直度",
            "卷形",
            "轧破甩尾",
        ]

    def readDataByDate(self, dateStr):
        sqlStatement = (
            "select * from unqual_histories where date(product_time) = '{}'"
            .format(dateStr)
        )
        df = pd.read_sql(sqlStatement, self.conn)
        print(df)
        return df


def main():

    dispatcher = UnqualHistoryDispatcher()
    # dispatcher.readDataByDate('2021-05-08')
    dispatcher.dispatchUnqualHistoryByDate('2021-05-20')


if __name__ == '__main__':
    main()
